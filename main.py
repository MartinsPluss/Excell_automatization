# datu nolasīšana no Excell faila, sakārtošana updatosana ar openpyxl
import openpyxl

# definē mainīgo lai atvērtu xls failu un konkrēto sheet1
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# izveido Dictionary prieks rezultatiem.

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10 = {}



# uztaisa for loopu lai zinātu cik reizes pārbaudīt katru rindu sakot ar rindu 2 līdz max row kas ir failā.

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value # norāda rindu kurā ir piegādātāja nosaukums 4 kolonna
    inventory = product_list.cell(product_row, 2).value # 2 kolonna
    price = product_list.cell(product_row, 3).value # 3 kolonna
    product_nr = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5) # write in new file!!!


    # 1) uzdevums: calculation number of products per supplier.

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name) # izmanto get funkciju, lai saņemtu datus no dictionary priekš key value
        products_per_supplier[supplier_name] = current_num_products + 1 # pieliek klāt 1 katru reizi kad iet cauri rindam
    else:
        products_per_supplier[supplier_name] = 1 # ja ir jauns piegādātājs nosaka ka tas ir pirmais ieraksts dictiorary

    # 2) uzdevums: calculation total value inventory per supplier

    if supplier_name in total_value_per_supplier:
            current_total_value = total_value_per_supplier.get(supplier_name)
            total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
            total_value_per_supplier[supplier_name] = inventory * price

    # 3) exercise list all products what are under 10 in inventory

    if inventory < 10:
        products_under_10[int(product_nr)] = int(inventory)



    # 4 exercise write summ of values of each product

    inventory_price.value = inventory * price


print(products_under_10)
print(products_per_supplier)
print(total_value_per_supplier)

inv_file.save("inventory_updated.xlsx")














